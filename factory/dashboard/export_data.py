# -*- coding: utf-8 -*-
"""
Export Data Module (Issues #255, #256)
======================================
Sistema de exportacao de dados em CSV, Excel e PDF.

Funcionalidades:
- Export de stories em CSV
- Export de stories em Excel (XLSX)
- Export de metricas em PDF
- Selecao de colunas
- Filtros aplicados no export
"""

from fastapi import APIRouter, Query, Response
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime
from io import BytesIO, StringIO
import csv
import json

from factory.database.connection import SessionLocal
from factory.database.models import Story, StoryTask, Sprint, Epic

router = APIRouter(prefix="/api/export", tags=["Export"])


@router.get("/stories/csv")
async def export_stories_csv(
    project_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    epic_id: Optional[str] = Query(None),
    sprint_id: Optional[str] = Query(None),
    columns: Optional[str] = Query(None, description="Colunas separadas por virgula")
):
    """Exporta stories em formato CSV."""
    db = SessionLocal()
    try:
        query = db.query(Story)

        if project_id:
            query = query.filter(Story.project_id == project_id)
        if status:
            query = query.filter(Story.status == status)
        if priority:
            query = query.filter(Story.priority == priority)
        if epic_id:
            query = query.filter(Story.epic_id == epic_id)
        if sprint_id:
            query = query.filter(Story.sprint_id == sprint_id)

        stories = query.order_by(Story.created_at.desc()).all()

        # Default columns
        default_cols = ['story_id', 'title', 'status', 'priority', 'story_points',
                        'assignee', 'epic_id', 'sprint_id', 'created_at', 'updated_at']
        selected_cols = columns.split(',') if columns else default_cols

        # Create CSV
        output = StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(selected_cols)

        # Data rows
        for story in stories:
            row = []
            for col in selected_cols:
                val = getattr(story, col, '')
                if hasattr(val, 'value'):
                    val = val.value
                elif isinstance(val, datetime):
                    val = val.strftime('%Y-%m-%d %H:%M:%S')
                row.append(val or '')
            writer.writerow(row)

        output.seek(0)
        filename = f"stories_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        db.close()


@router.get("/stories/excel")
async def export_stories_excel(
    project_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    epic_id: Optional[str] = Query(None),
    sprint_id: Optional[str] = Query(None)
):
    """Exporta stories em formato Excel (XLSX)."""
    db = SessionLocal()
    try:
        # Try to import openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        query = db.query(Story)

        if project_id:
            query = query.filter(Story.project_id == project_id)
        if status:
            query = query.filter(Story.status == status)
        if priority:
            query = query.filter(Story.priority == priority)
        if epic_id:
            query = query.filter(Story.epic_id == epic_id)
        if sprint_id:
            query = query.filter(Story.sprint_id == sprint_id)

        stories = query.order_by(Story.created_at.desc()).all()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Stories"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003B4A", end_color="003B4A", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = ['ID', 'Titulo', 'Status', 'Prioridade', 'Pontos',
                   'Responsavel', 'Epic', 'Sprint', 'Criado em', 'Atualizado em']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data
        for row_num, story in enumerate(stories, 2):
            data = [
                story.story_id,
                story.title,
                story.status.value if hasattr(story.status, 'value') else story.status,
                story.priority.value if hasattr(story.priority, 'value') else story.priority,
                story.story_points,
                story.assignee or '',
                story.epic_id or '',
                story.sprint_id or '',
                story.created_at.strftime('%Y-%m-%d %H:%M') if story.created_at else '',
                story.updated_at.strftime('%Y-%m-%d %H:%M') if story.updated_at else ''
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"stories_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        db.close()


@router.get("/tasks/csv")
async def export_tasks_csv(story_id: Optional[str] = Query(None)):
    """Exporta tasks em formato CSV."""
    db = SessionLocal()
    try:
        query = db.query(StoryTask)
        if story_id:
            query = query.filter(StoryTask.story_id == story_id)

        tasks = query.order_by(StoryTask.created_at.desc()).all()

        output = StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(['task_id', 'story_id', 'title', 'status', 'task_type',
                        'progress', 'estimated_hours', 'assignee', 'created_at'])

        # Data
        for task in tasks:
            writer.writerow([
                task.task_id,
                task.story_id,
                task.title,
                task.status.value if hasattr(task.status, 'value') else task.status,
                task.task_type.value if hasattr(task.task_type, 'value') else task.task_type,
                task.progress,
                getattr(task, 'estimated_hours', 0),
                getattr(task, 'assignee', ''),
                task.created_at.strftime('%Y-%m-%d %H:%M:%S') if task.created_at else ''
            ])

        output.seek(0)
        filename = f"tasks_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        db.close()


@router.get("/metrics/csv")
async def export_metrics_csv(
    project_id: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
):
    """Exporta metricas agregadas em CSV."""
    db = SessionLocal()
    try:
        query = db.query(Story)
        if project_id:
            query = query.filter(Story.project_id == project_id)

        stories = query.all()

        # Calculate metrics
        total_stories = len(stories)
        by_status = {}
        by_priority = {}
        total_points = 0

        for story in stories:
            status = story.status.value if hasattr(story.status, 'value') else story.status
            priority = story.priority.value if hasattr(story.priority, 'value') else story.priority

            by_status[status] = by_status.get(status, 0) + 1
            by_priority[priority] = by_priority.get(priority, 0) + 1
            total_points += story.story_points or 0

        output = StringIO()
        writer = csv.writer(output)

        writer.writerow(['Metrica', 'Valor'])
        writer.writerow(['Total de Stories', total_stories])
        writer.writerow(['Total de Pontos', total_points])
        writer.writerow([''])
        writer.writerow(['Por Status', ''])
        for status, count in by_status.items():
            writer.writerow([f'  {status}', count])
        writer.writerow([''])
        writer.writerow(['Por Prioridade', ''])
        for priority, count in by_priority.items():
            writer.writerow([f'  {priority}', count])

        output.seek(0)
        filename = f"metrics_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        db.close()


@router.get("/sprint/{sprint_id}/csv")
async def export_sprint_csv(sprint_id: str):
    """Exporta dados de um sprint em CSV."""
    db = SessionLocal()
    try:
        sprint = db.query(Sprint).filter(Sprint.sprint_id == sprint_id).first()
        if not sprint:
            return {"error": "Sprint nao encontrado"}

        stories = db.query(Story).filter(Story.sprint_id == sprint_id).all()

        output = StringIO()
        writer = csv.writer(output)

        # Sprint info
        writer.writerow(['Sprint', sprint.name])
        writer.writerow(['Goal', sprint.goal or ''])
        writer.writerow(['Inicio', sprint.start_date.strftime('%Y-%m-%d') if sprint.start_date else ''])
        writer.writerow(['Fim', sprint.end_date.strftime('%Y-%m-%d') if sprint.end_date else ''])
        writer.writerow([''])
        writer.writerow(['Stories do Sprint'])
        writer.writerow(['ID', 'Titulo', 'Status', 'Pontos', 'Responsavel'])

        total_points = 0
        completed_points = 0
        for story in stories:
            status = story.status.value if hasattr(story.status, 'value') else story.status
            writer.writerow([
                story.story_id,
                story.title,
                status,
                story.story_points,
                story.assignee or ''
            ])
            total_points += story.story_points or 0
            if status == 'done':
                completed_points += story.story_points or 0

        writer.writerow([''])
        writer.writerow(['Total de Pontos', total_points])
        writer.writerow(['Pontos Completados', completed_points])
        writer.writerow(['Velocidade', f"{(completed_points/total_points*100):.1f}%" if total_points > 0 else '0%'])

        output.seek(0)
        filename = f"sprint_{sprint_id}_{datetime.now().strftime('%Y%m%d')}.csv"

        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        db.close()


def register_export_data(app):
    """Registra os endpoints de export no app FastAPI."""
    app.include_router(router)
    print("[Dashboard] Export Data endpoints loaded: /api/export/*")
