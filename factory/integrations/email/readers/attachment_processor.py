# -*- coding: utf-8 -*-
"""
Attachment Processor
====================
Processador de anexos de emails.

Funcionalidades:
- Baixar anexos de emails
- Processar documentos PDF, Word, Excel
- Extrair texto e requisitos
- Salvar anexos em diretorio de projeto

Uso:
    from factory.integrations.email.readers import AttachmentProcessor

    processor = AttachmentProcessor(graph_client)
    attachments = await processor.get_attachments(email.id)

    for att in attachments:
        content = await processor.extract_text(att)
        print(content)
"""

import os
import io
import re
import logging
import mimetypes
from pathlib import Path
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Union
from enum import Enum

logger = logging.getLogger(__name__)


class AttachmentType(str, Enum):
    """Tipo de anexo"""
    PDF = "pdf"
    WORD = "word"
    EXCEL = "excel"
    POWERPOINT = "powerpoint"
    IMAGE = "image"
    TEXT = "text"
    CODE = "code"
    ARCHIVE = "archive"
    OTHER = "other"


@dataclass
class ProcessedAttachment:
    """Anexo processado com metadados"""
    id: str
    name: str
    content_type: str
    size: int
    attachment_type: AttachmentType
    content: Optional[bytes] = None
    extracted_text: Optional[str] = None
    extracted_data: Dict[str, Any] = field(default_factory=dict)
    saved_path: Optional[Path] = None

    @property
    def extension(self) -> str:
        """Retorna extensao do arquivo"""
        return Path(self.name).suffix.lower()

    @property
    def size_formatted(self) -> str:
        """Retorna tamanho formatado"""
        if self.size < 1024:
            return f"{self.size} B"
        elif self.size < 1024 * 1024:
            return f"{self.size / 1024:.1f} KB"
        else:
            return f"{self.size / (1024 * 1024):.1f} MB"


class AttachmentProcessor:
    """
    Processador de anexos de emails.

    Suporta:
    - PDF (com PyPDF2 ou pdfplumber)
    - Word (.docx com python-docx)
    - Excel (.xlsx com openpyxl)
    - Texto/Codigo (utf-8)
    - Imagens (metadados)

    Exemplo:
        client = MicrosoftGraphClient(config)
        processor = AttachmentProcessor(client)

        # Baixar e processar anexos
        attachments = await processor.get_attachments(message_id)

        for att in attachments:
            if att.attachment_type == AttachmentType.PDF:
                text = await processor.extract_text(att)
                requirements = processor.extract_requirements(text)
    """

    # Mapeamento de MIME types para tipos
    MIME_TYPE_MAPPING = {
        "application/pdf": AttachmentType.PDF,
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": AttachmentType.WORD,
        "application/msword": AttachmentType.WORD,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": AttachmentType.EXCEL,
        "application/vnd.ms-excel": AttachmentType.EXCEL,
        "application/vnd.openxmlformats-officedocument.presentationml.presentation": AttachmentType.POWERPOINT,
        "application/vnd.ms-powerpoint": AttachmentType.POWERPOINT,
        "text/plain": AttachmentType.TEXT,
        "text/csv": AttachmentType.TEXT,
        "application/json": AttachmentType.CODE,
        "text/html": AttachmentType.CODE,
        "text/xml": AttachmentType.CODE,
        "application/zip": AttachmentType.ARCHIVE,
        "application/x-rar-compressed": AttachmentType.ARCHIVE,
        "application/x-7z-compressed": AttachmentType.ARCHIVE
    }

    # Extensoes de codigo
    CODE_EXTENSIONS = {
        ".py", ".js", ".ts", ".jsx", ".tsx", ".java", ".cs", ".cpp",
        ".c", ".h", ".go", ".rs", ".rb", ".php", ".swift", ".kt",
        ".json", ".xml", ".yaml", ".yml", ".md", ".sql"
    }

    # Extensoes de imagem
    IMAGE_EXTENSIONS = {
        ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".svg", ".webp", ".ico"
    }

    def __init__(
        self,
        graph_client,
        upload_dir: Optional[Path] = None,
        max_size_mb: int = 25
    ):
        """
        Inicializa o processador.

        Args:
            graph_client: Cliente Microsoft Graph
            upload_dir: Diretorio para salvar arquivos
            max_size_mb: Tamanho maximo em MB
        """
        self.client = graph_client
        self.upload_dir = upload_dir or Path("uploads")
        self.max_size_bytes = max_size_mb * 1024 * 1024
        self._ensure_upload_dir()

    def _ensure_upload_dir(self):
        """Garante que diretorio de upload existe"""
        self.upload_dir.mkdir(parents=True, exist_ok=True)

    def _detect_type(self, name: str, content_type: str) -> AttachmentType:
        """Detecta tipo do anexo"""
        # Primeiro tenta pelo MIME type
        if content_type in self.MIME_TYPE_MAPPING:
            return self.MIME_TYPE_MAPPING[content_type]

        # Depois pela extensao
        ext = Path(name).suffix.lower()

        if ext in self.CODE_EXTENSIONS:
            return AttachmentType.CODE

        if ext in self.IMAGE_EXTENSIONS:
            return AttachmentType.IMAGE

        if ext == ".pdf":
            return AttachmentType.PDF

        if ext in (".doc", ".docx"):
            return AttachmentType.WORD

        if ext in (".xls", ".xlsx"):
            return AttachmentType.EXCEL

        if ext in (".ppt", ".pptx"):
            return AttachmentType.POWERPOINT

        if ext == ".txt":
            return AttachmentType.TEXT

        if ext in (".zip", ".rar", ".7z", ".tar", ".gz"):
            return AttachmentType.ARCHIVE

        return AttachmentType.OTHER

    async def get_attachments(
        self,
        message_id: str,
        download_content: bool = True
    ) -> List[ProcessedAttachment]:
        """
        Busca e processa anexos de uma mensagem.

        Args:
            message_id: ID da mensagem
            download_content: Se deve baixar o conteudo

        Returns:
            List[ProcessedAttachment]
        """
        try:
            attachments_data = await self.client.get_message_attachments(message_id)
            processed = []

            for att_data in attachments_data:
                att = self._process_attachment_metadata(att_data)

                if download_content and att.size <= self.max_size_bytes:
                    content = await self.client.download_attachment(
                        message_id,
                        att.id
                    )
                    att.content = content

                processed.append(att)

            return processed

        except Exception as e:
            logger.exception(f"Erro ao processar anexos: {e}")
            return []

    def _process_attachment_metadata(
        self,
        att_data: Dict[str, Any]
    ) -> ProcessedAttachment:
        """Processa metadados do anexo"""
        name = att_data.get("name", "attachment")
        content_type = att_data.get("contentType", "application/octet-stream")
        size = att_data.get("size", 0)

        return ProcessedAttachment(
            id=att_data.get("id", ""),
            name=name,
            content_type=content_type,
            size=size,
            attachment_type=self._detect_type(name, content_type)
        )

    async def extract_text(
        self,
        attachment: ProcessedAttachment
    ) -> Optional[str]:
        """
        Extrai texto de um anexo.

        Args:
            attachment: Anexo a processar

        Returns:
            str ou None
        """
        if not attachment.content:
            logger.warning(f"Anexo '{attachment.name}' sem conteudo")
            return None

        try:
            if attachment.attachment_type == AttachmentType.PDF:
                return self._extract_text_from_pdf(attachment.content)

            elif attachment.attachment_type == AttachmentType.WORD:
                return self._extract_text_from_docx(attachment.content)

            elif attachment.attachment_type == AttachmentType.EXCEL:
                return self._extract_text_from_xlsx(attachment.content)

            elif attachment.attachment_type in (AttachmentType.TEXT, AttachmentType.CODE):
                return self._extract_text_from_text(attachment.content)

            else:
                logger.info(f"Tipo '{attachment.attachment_type}' nao suporta extracao de texto")
                return None

        except Exception as e:
            logger.error(f"Erro ao extrair texto de '{attachment.name}': {e}")
            return None

    def _extract_text_from_pdf(self, content: bytes) -> str:
        """Extrai texto de PDF"""
        try:
            # Tenta pdfplumber (melhor qualidade)
            import pdfplumber

            with pdfplumber.open(io.BytesIO(content)) as pdf:
                text_parts = []
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
                return "\n\n".join(text_parts)

        except ImportError:
            pass

        try:
            # Fallback para PyPDF2
            from PyPDF2 import PdfReader

            reader = PdfReader(io.BytesIO(content))
            text_parts = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
            return "\n\n".join(text_parts)

        except ImportError:
            logger.warning("Nenhuma biblioteca PDF disponivel. Instale: pdfplumber ou PyPDF2")
            return ""

    def _extract_text_from_docx(self, content: bytes) -> str:
        """Extrai texto de DOCX"""
        try:
            from docx import Document

            doc = Document(io.BytesIO(content))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs)

        except ImportError:
            logger.warning("python-docx nao instalado. Execute: pip install python-docx")
            return ""

    def _extract_text_from_xlsx(self, content: bytes) -> str:
        """Extrai texto de XLSX"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(content), data_only=True)
            text_parts = []

            for sheet in wb.worksheets:
                text_parts.append(f"=== {sheet.title} ===")
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_values):
                        text_parts.append(" | ".join(row_values))

            return "\n".join(text_parts)

        except ImportError:
            logger.warning("openpyxl nao instalado. Execute: pip install openpyxl")
            return ""

    def _extract_text_from_text(self, content: bytes) -> str:
        """Extrai texto de arquivos texto/codigo"""
        encodings = ["utf-8", "latin-1", "cp1252"]

        for encoding in encodings:
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue

        return content.decode("utf-8", errors="ignore")

    def extract_requirements(self, text: str) -> List[str]:
        """
        Extrai requisitos de texto.

        Args:
            text: Texto a analisar

        Returns:
            List[str]: Lista de requisitos
        """
        requirements = []

        if not text:
            return requirements

        lines = text.split("\n")
        in_req_section = False

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Detecta secao de requisitos
            if re.match(r"^(requisitos?|requirements?|funcionalidades?|features?|especifica.+es?)[\s:]*$", line, re.IGNORECASE):
                in_req_section = True
                continue

            # Detecta fim de secao
            if in_req_section and re.match(r"^[A-Z][^:]*:$", line):
                in_req_section = False

            # Items de lista
            match = re.match(r"^[\d\.\-\*\+•○►]+\s*(.{10,})", line)
            if match:
                req = match.group(1).strip()
                # Limpa caracteres extras
                req = re.sub(r"\s+", " ", req)
                if len(req) > 10 and len(req) < 500:
                    requirements.append(req)
            elif in_req_section and len(line) > 15:
                requirements.append(line)

        return requirements[:50]  # Limita a 50 requisitos

    def extract_user_stories(self, text: str) -> List[Dict[str, str]]:
        """
        Extrai user stories do texto.

        Args:
            text: Texto a analisar

        Returns:
            List[Dict]: Lista de user stories
        """
        stories = []

        if not text:
            return stories

        # Padrao: Como um [persona], eu quero [acao], para [beneficio]
        pattern = r"[Cc]omo\s+(?:um|uma|o|a)\s+(.+?),?\s+(?:eu\s+)?quero\s+(.+?),?\s+para\s+(?:que\s+)?(.+?)(?:\.|$)"

        matches = re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE)

        for match in matches:
            stories.append({
                "persona": match.group(1).strip(),
                "action": match.group(2).strip(),
                "benefit": match.group(3).strip()
            })

        return stories

    async def save_attachment(
        self,
        attachment: ProcessedAttachment,
        project_dir: Optional[Path] = None,
        subdir: str = "attachments"
    ) -> Optional[Path]:
        """
        Salva anexo em disco.

        Args:
            attachment: Anexo a salvar
            project_dir: Diretorio do projeto
            subdir: Subdiretorio

        Returns:
            Path do arquivo salvo ou None
        """
        if not attachment.content:
            logger.warning(f"Anexo '{attachment.name}' sem conteudo para salvar")
            return None

        try:
            # Define diretorio de destino
            if project_dir:
                target_dir = project_dir / subdir
            else:
                target_dir = self.upload_dir / subdir

            target_dir.mkdir(parents=True, exist_ok=True)

            # Gera nome unico se necessario
            file_path = target_dir / attachment.name
            if file_path.exists():
                stem = file_path.stem
                suffix = file_path.suffix
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_path = target_dir / f"{stem}_{timestamp}{suffix}"

            # Salva arquivo
            with open(file_path, "wb") as f:
                f.write(attachment.content)

            attachment.saved_path = file_path
            logger.info(f"Anexo salvo: {file_path}")
            return file_path

        except Exception as e:
            logger.error(f"Erro ao salvar anexo '{attachment.name}': {e}")
            return None

    async def process_email_attachments(
        self,
        message_id: str,
        project_dir: Optional[Path] = None,
        extract_text: bool = True,
        save_files: bool = True
    ) -> Dict[str, Any]:
        """
        Processa todos os anexos de um email.

        Args:
            message_id: ID da mensagem
            project_dir: Diretorio do projeto
            extract_text: Se deve extrair texto
            save_files: Se deve salvar arquivos

        Returns:
            Dict com resultados do processamento
        """
        result = {
            "attachments": [],
            "requirements": [],
            "user_stories": [],
            "saved_files": [],
            "errors": []
        }

        try:
            attachments = await self.get_attachments(message_id)

            for att in attachments:
                att_info = {
                    "name": att.name,
                    "type": att.attachment_type.value,
                    "size": att.size_formatted
                }

                # Extrai texto
                if extract_text:
                    text = await self.extract_text(att)
                    if text:
                        att.extracted_text = text

                        # Extrai requisitos
                        reqs = self.extract_requirements(text)
                        result["requirements"].extend(reqs)

                        # Extrai user stories
                        stories = self.extract_user_stories(text)
                        result["user_stories"].extend(stories)

                # Salva arquivo
                if save_files and att.content:
                    saved_path = await self.save_attachment(att, project_dir)
                    if saved_path:
                        att_info["saved_path"] = str(saved_path)
                        result["saved_files"].append(str(saved_path))

                result["attachments"].append(att_info)

        except Exception as e:
            logger.exception(f"Erro ao processar anexos: {e}")
            result["errors"].append(str(e))

        # Remove duplicatas de requisitos
        result["requirements"] = list(dict.fromkeys(result["requirements"]))

        return result
