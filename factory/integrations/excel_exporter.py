# -*- coding: utf-8 -*-
"""
Excel Exporter Module
=====================
Exportador de dados para Excel com suporte multi-tenant.

Funcionalidades:
- Exportar projetos completos para Excel
- Multiplas planilhas (Stories, Tasks, Metricas)
- Formatacao profissional
- Graficos e dashboards
- Isolamento por tenant

Dependencias:
- openpyxl: Para criar arquivos .xlsx
"""

import io
import logging
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, BinaryIO
from enum import Enum

logger = logging.getLogger(__name__)

# Tentar importar openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("[Excel] openpyxl nao instalado. Instale com: pip install openpyxl")


class ExportFormat(str, Enum):
    """Formatos de exportacao"""
    XLSX = "xlsx"
    CSV = "csv"


@dataclass
class ExcelStyle:
    """Estilos para formatacao Excel"""
    # Cores Belgo Arames
    header_fill: str = "003B4A"  # Azul Belgo
    header_font: str = "FFFFFF"
    accent_fill: str = "FF6C00"  # Laranja Belgo
    success_fill: str = "10B981"  # Verde
    warning_fill: str = "F59E0B"  # Amarelo
    error_fill: str = "EF4444"   # Vermelho

    # Fontes
    header_font_size: int = 12
    data_font_size: int = 10
    title_font_size: int = 16


class ExcelExporter:
    """
    Exportador de projetos para Excel.

    Gera arquivos Excel com multiplas planilhas contendo
    todas as informacoes do projeto, tasks e metricas.

    Exemplo de uso:
    ```python
    exporter = ExcelExporter(tenant_id="...")

    # Exportar projeto completo
    excel_bytes = exporter.export_project(
        project_id="PROJ-001",
        include_tasks=True,
        include_metrics=True
    )

    # Salvar arquivo
    with open("projeto.xlsx", "wb") as f:
        f.write(excel_bytes)
    ```
    """

    def __init__(self, tenant_id: str, style: Optional[ExcelStyle] = None):
        """
        Inicializa o exportador Excel.

        Args:
            tenant_id: ID do tenant para isolamento de dados
            style: Estilos customizados (opcional)
        """
        self.tenant_id = tenant_id
        self.style = style or ExcelStyle()
        self._rate_limit_remaining = 50
        self._rate_limit_reset: Optional[datetime] = None

        if not OPENPYXL_AVAILABLE:
            logger.error("[Excel] openpyxl nao disponivel. Exportacao desabilitada.")

    def _check_rate_limit(self) -> bool:
        """Verifica se esta dentro do rate limit"""
        if self._rate_limit_reset and datetime.utcnow() > self._rate_limit_reset:
            self._rate_limit_remaining = 50
            self._rate_limit_reset = None

        if self._rate_limit_remaining <= 0:
            logger.warning(f"[Excel] Rate limit atingido para tenant {self.tenant_id}")
            return False

        self._rate_limit_remaining -= 1
        return True

    def _log_operation(self, operation: str, details: Dict[str, Any] = None):
        """Registra operacao para auditoria"""
        logger.info(f"[Excel] {operation} - tenant: {self.tenant_id}, details: {details}")

    def _create_header_style(self, wb: 'Workbook') -> Dict:
        """Cria estilos para cabecalhos"""
        return {
            "font": Font(
                name='Calibri',
                size=self.style.header_font_size,
                bold=True,
                color=self.style.header_font
            ),
            "fill": PatternFill(
                start_color=self.style.header_fill,
                end_color=self.style.header_fill,
                fill_type="solid"
            ),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def _apply_header_style(self, cell, styles: Dict):
        """Aplica estilos ao cabecalho"""
        cell.font = styles["font"]
        cell.fill = styles["fill"]
        cell.alignment = styles["alignment"]
        cell.border = styles["border"]

    def _get_status_fill(self, status: str) -> 'PatternFill':
        """Retorna preenchimento baseado no status"""
        status_colors = {
            "done": self.style.success_fill,
            "in_progress": self.style.accent_fill,
            "testing": self.style.warning_fill,
            "blocked": self.style.error_fill,
            "review": "9333EA",  # Roxo
            "backlog": "6B7280",  # Cinza
            "ready": "3B82F6"    # Azul
        }
        color = status_colors.get(status, "FFFFFF")
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def export_project(
        self,
        project_id: str,
        stories: Optional[List[Dict]] = None,
        include_tasks: bool = True,
        include_metrics: bool = True,
        include_charts: bool = True
    ) -> bytes:
        """
        Exportar projeto completo para Excel.

        Args:
            project_id: ID do projeto
            stories: Lista de stories (se None, busca do banco)
            include_tasks: Incluir planilha de tasks
            include_metrics: Incluir planilha de metricas
            include_charts: Incluir graficos

        Returns:
            bytes: Conteudo do arquivo Excel
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl nao instalado. Instale com: pip install openpyxl")

        if not self._check_rate_limit():
            raise RuntimeError("Rate limit exceeded. Aguarde antes de tentar novamente.")

        self._log_operation("export_project", {
            "project_id": project_id,
            "include_tasks": include_tasks,
            "include_metrics": include_metrics
        })

        # Buscar stories se nao fornecidas
        if stories is None:
            stories = self._get_stories_from_db(project_id)

        # Filtrar por tenant
        stories = [s for s in stories if self._is_tenant_authorized(s)]

        # Criar workbook
        wb = Workbook()
        header_styles = self._create_header_style(wb)

        # Planilha de Stories (ativa por padrao)
        ws_stories = wb.active
        ws_stories.title = "Funcionalidades"
        self._create_stories_sheet(ws_stories, stories, header_styles)

        # Planilha de Tasks
        if include_tasks:
            ws_tasks = wb.create_sheet("Tarefas")
            tasks = self._extract_tasks(stories)
            self._create_tasks_sheet(ws_tasks, tasks, header_styles)

        # Planilha de Metricas
        if include_metrics:
            ws_metrics = wb.create_sheet("Metricas")
            self._create_metrics_sheet(ws_metrics, stories, header_styles)

        # Dashboard com graficos
        if include_charts:
            ws_dashboard = wb.create_sheet("Dashboard", 0)  # Primeira aba
            self._create_dashboard_sheet(ws_dashboard, stories, header_styles)

        # Planilha de Resumo
        ws_summary = wb.create_sheet("Resumo", 1)
        self._create_summary_sheet(ws_summary, project_id, stories, header_styles)

        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    def _create_stories_sheet(
        self,
        ws,
        stories: List[Dict],
        header_styles: Dict
    ):
        """Cria planilha de stories"""
        # Headers
        headers = [
            "ID", "Titulo", "Status", "Prioridade", "Pontos",
            "Complexidade", "Categoria", "Epico", "Sprint",
            "Responsavel", "Progresso", "Tasks", "Criado em"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell, header_styles)

        # Dados
        for row, story in enumerate(stories, 2):
            ws.cell(row=row, column=1, value=story.get("story_id", ""))
            ws.cell(row=row, column=2, value=story.get("title", ""))

            status_cell = ws.cell(row=row, column=3, value=story.get("status", ""))
            status_cell.fill = self._get_status_fill(story.get("status", ""))

            ws.cell(row=row, column=4, value=story.get("priority", ""))
            ws.cell(row=row, column=5, value=story.get("story_points", 0))
            ws.cell(row=row, column=6, value=story.get("complexity", ""))
            ws.cell(row=row, column=7, value=story.get("category", ""))
            ws.cell(row=row, column=8, value=story.get("epic_id", ""))
            ws.cell(row=row, column=9, value=story.get("sprint_id", ""))
            ws.cell(row=row, column=10, value=story.get("assignee", ""))

            progress = story.get("progress", 0)
            progress_cell = ws.cell(row=row, column=11, value=f"{progress}%")
            progress_cell.alignment = Alignment(horizontal="center")

            tasks_total = story.get("tasks_total", 0)
            tasks_completed = story.get("tasks_completed", 0)
            ws.cell(row=row, column=12, value=f"{tasks_completed}/{tasks_total}")

            created_at = story.get("created_at")
            if isinstance(created_at, str):
                ws.cell(row=row, column=13, value=created_at[:10])
            elif created_at:
                ws.cell(row=row, column=13, value=created_at.strftime("%Y-%m-%d"))

        # Ajustar largura das colunas
        self._auto_adjust_columns(ws)

        # Adicionar tabela formatada
        if len(stories) > 0:
            table_range = f"A1:{get_column_letter(len(headers))}{len(stories) + 1}"
            table = Table(displayName="TabelaFuncionalidades", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)

    def _create_tasks_sheet(
        self,
        ws,
        tasks: List[Dict],
        header_styles: Dict
    ):
        """Cria planilha de tasks"""
        headers = [
            "ID", "Story", "Titulo", "Tipo", "Status",
            "Progresso", "Horas Est.", "Horas Real",
            "Responsavel", "Agente IA"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell, header_styles)

        for row, task in enumerate(tasks, 2):
            ws.cell(row=row, column=1, value=task.get("task_id", ""))
            ws.cell(row=row, column=2, value=task.get("story_id", ""))
            ws.cell(row=row, column=3, value=task.get("title", ""))
            ws.cell(row=row, column=4, value=task.get("task_type", ""))

            status_cell = ws.cell(row=row, column=5, value=task.get("status", ""))
            status_cell.fill = self._get_status_fill(task.get("status", ""))

            ws.cell(row=row, column=6, value=f"{task.get('progress', 0)}%")
            ws.cell(row=row, column=7, value=task.get("estimated_hours", 0))
            ws.cell(row=row, column=8, value=task.get("actual_hours", 0))
            ws.cell(row=row, column=9, value=task.get("assignee", ""))
            ws.cell(row=row, column=10, value=task.get("agent_id", ""))

        self._auto_adjust_columns(ws)

    def _create_metrics_sheet(
        self,
        ws,
        stories: List[Dict],
        header_styles: Dict
    ):
        """Cria planilha de metricas"""
        # Calcular metricas
        total = len(stories)
        done = len([s for s in stories if s.get("status") == "done"])
        in_progress = len([s for s in stories if s.get("status") == "in_progress"])
        total_points = sum(s.get("story_points", 0) for s in stories)
        done_points = sum(s.get("story_points", 0) for s in stories if s.get("status") == "done")

        # Cabecalho
        title_cell = ws.cell(row=1, column=1, value="Metricas do Projeto")
        title_cell.font = Font(size=self.style.title_font_size, bold=True)
        ws.merge_cells("A1:D1")

        # Data de geracao
        ws.cell(row=2, column=1, value=f"Gerado em: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}")

        # Metricas
        metrics_data = [
            ("Total de Funcionalidades", total),
            ("Funcionalidades Concluidas", done),
            ("Em Progresso", in_progress),
            ("Taxa de Conclusao", f"{(done/total*100) if total > 0 else 0:.1f}%"),
            ("", ""),
            ("Total de Story Points", total_points),
            ("Points Entregues", done_points),
            ("Velocidade Media (estimada)", f"{done_points / 2:.1f} pts/sprint"),
        ]

        for row, (label, value) in enumerate(metrics_data, 4):
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

        # Status distribution
        ws.cell(row=14, column=1, value="Distribuicao por Status")
        ws.cell(row=14, column=1).font = Font(bold=True)

        status_counts = {}
        for story in stories:
            status = story.get("status", "backlog")
            status_counts[status] = status_counts.get(status, 0) + 1

        for row, (status, count) in enumerate(status_counts.items(), 15):
            ws.cell(row=row, column=1, value=status.replace("_", " ").title())
            ws.cell(row=row, column=2, value=count)

        self._auto_adjust_columns(ws)

    def _create_dashboard_sheet(
        self,
        ws,
        stories: List[Dict],
        header_styles: Dict
    ):
        """Cria dashboard com graficos"""
        # Titulo
        title_cell = ws.cell(row=1, column=1, value="Dashboard do Projeto")
        title_cell.font = Font(size=self.style.title_font_size, bold=True, color=self.style.header_fill)
        ws.merge_cells("A1:H1")

        # KPIs
        total = len(stories)
        done = len([s for s in stories if s.get("status") == "done"])
        total_points = sum(s.get("story_points", 0) for s in stories)
        done_points = sum(s.get("story_points", 0) for s in stories if s.get("status") == "done")

        kpis = [
            ("Total", total),
            ("Concluidas", done),
            ("Em Progresso", len([s for s in stories if s.get("status") == "in_progress"])),
            ("Story Points", total_points),
            ("Points Entregues", done_points),
            ("% Conclusao", f"{(done/total*100) if total > 0 else 0:.0f}%")
        ]

        for col, (label, value) in enumerate(kpis, 1):
            label_cell = ws.cell(row=3, column=col, value=label)
            label_cell.font = Font(bold=True, size=10)
            label_cell.alignment = Alignment(horizontal="center")

            value_cell = ws.cell(row=4, column=col, value=value)
            value_cell.font = Font(bold=True, size=14)
            value_cell.alignment = Alignment(horizontal="center")
            if col <= 3:
                value_cell.fill = PatternFill(
                    start_color=self.style.accent_fill,
                    end_color=self.style.accent_fill,
                    fill_type="solid"
                )
                value_cell.font = Font(bold=True, size=14, color="FFFFFF")

        # Dados para graficos (status distribution)
        ws.cell(row=7, column=1, value="Status")
        ws.cell(row=7, column=2, value="Quantidade")

        status_order = ["backlog", "ready", "in_progress", "review", "testing", "done"]
        status_counts = {s: 0 for s in status_order}
        for story in stories:
            status = story.get("status", "backlog")
            if status in status_counts:
                status_counts[status] += 1

        for row, status in enumerate(status_order, 8):
            ws.cell(row=row, column=1, value=status.replace("_", " ").title())
            ws.cell(row=row, column=2, value=status_counts[status])

        # Grafico de barras - Status
        if len(stories) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Funcionalidades por Status"
            chart.y_axis.title = "Quantidade"

            data = Reference(ws, min_col=2, min_row=7, max_row=13)
            cats = Reference(ws, min_col=1, min_row=8, max_row=13)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4

            ws.add_chart(chart, "D7")

            # Grafico de pizza - Prioridade
            ws.cell(row=16, column=1, value="Prioridade")
            ws.cell(row=16, column=2, value="Quantidade")

            priority_counts = {}
            for story in stories:
                priority = story.get("priority", "medium")
                priority_counts[priority] = priority_counts.get(priority, 0) + 1

            for row, (priority, count) in enumerate(priority_counts.items(), 17):
                ws.cell(row=row, column=1, value=priority.title())
                ws.cell(row=row, column=2, value=count)

            pie = PieChart()
            pie.title = "Distribuicao por Prioridade"
            labels = Reference(ws, min_col=1, min_row=17, max_row=16 + len(priority_counts))
            data = Reference(ws, min_col=2, min_row=16, max_row=16 + len(priority_counts))
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)

            ws.add_chart(pie, "D18")

        self._auto_adjust_columns(ws)

    def _create_summary_sheet(
        self,
        ws,
        project_id: str,
        stories: List[Dict],
        header_styles: Dict
    ):
        """Cria planilha de resumo"""
        ws.cell(row=1, column=1, value="Resumo do Projeto")
        ws.cell(row=1, column=1).font = Font(size=self.style.title_font_size, bold=True)

        info = [
            ("Projeto", project_id),
            ("Tenant", self.tenant_id),
            ("Data de Exportacao", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("Total de Funcionalidades", len(stories)),
            ("Total de Story Points", sum(s.get("story_points", 0) for s in stories)),
            ("", ""),
            ("Exportado por", "Fabrica de Agentes"),
            ("Versao", "6.5")
        ]

        for row, (label, value) in enumerate(info, 3):
            if label:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

        self._auto_adjust_columns(ws)

    def _extract_tasks(self, stories: List[Dict]) -> List[Dict]:
        """Extrai todas as tasks das stories"""
        tasks = []
        for story in stories:
            story_tasks = story.get("tasks", [])
            if story_tasks:
                for task in story_tasks:
                    task["story_id"] = story.get("story_id")
                    tasks.append(task)
        return tasks

    def _auto_adjust_columns(self, ws):
        """Ajusta largura das colunas automaticamente"""
        for column_cells in ws.columns:
            max_length = 0
            column = None
            for cell in column_cells:
                if column is None:
                    column = cell.column_letter
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            if column:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

    def _is_tenant_authorized(self, resource: Dict) -> bool:
        """Verifica se o recurso pertence ao tenant"""
        resource_tenant = resource.get("tenant_id")
        if resource_tenant is None:
            return True
        return resource_tenant == self.tenant_id

    def _get_stories_from_db(self, project_id: str) -> List[Dict]:
        """Busca stories do banco de dados"""
        try:
            from factory.database.connection import get_session
            from factory.database.models import Story, StoryTask

            with get_session() as session:
                stories = session.query(Story).filter(
                    Story.project_id == project_id
                ).all()

                result = []
                for story in stories:
                    story_dict = story.to_dict()
                    # Buscar tasks
                    tasks = session.query(StoryTask).filter(
                        StoryTask.story_id == story.story_id
                    ).all()
                    story_dict["tasks"] = [t.to_dict() for t in tasks]
                    result.append(story_dict)

                return result
        except Exception as e:
            logger.error(f"[Excel] Erro ao buscar stories: {e}")
            return []

    def export_to_csv(
        self,
        project_id: str,
        stories: Optional[List[Dict]] = None
    ) -> str:
        """
        Exportar stories para CSV.

        Args:
            project_id: ID do projeto
            stories: Lista de stories (opcional)

        Returns:
            str: Conteudo CSV
        """
        import csv
        from io import StringIO

        if not self._check_rate_limit():
            raise RuntimeError("Rate limit exceeded")

        if stories is None:
            stories = self._get_stories_from_db(project_id)

        stories = [s for s in stories if self._is_tenant_authorized(s)]

        output = StringIO()
        writer = csv.writer(output)

        # Header
        headers = [
            "ID", "Titulo", "Status", "Prioridade", "Pontos",
            "Complexidade", "Categoria", "Epico", "Sprint",
            "Responsavel", "Progresso", "Criado em"
        ]
        writer.writerow(headers)

        # Dados
        for story in stories:
            writer.writerow([
                story.get("story_id", ""),
                story.get("title", ""),
                story.get("status", ""),
                story.get("priority", ""),
                story.get("story_points", 0),
                story.get("complexity", ""),
                story.get("category", ""),
                story.get("epic_id", ""),
                story.get("sprint_id", ""),
                story.get("assignee", ""),
                story.get("progress", 0),
                story.get("created_at", "")[:10] if story.get("created_at") else ""
            ])

        return output.getvalue()


# Cache de exporters por tenant
_exporters: Dict[str, ExcelExporter] = {}


def get_excel_exporter(tenant_id: str) -> ExcelExporter:
    """
    Retorna instancia do exportador Excel para o tenant.

    Args:
        tenant_id: ID do tenant

    Returns:
        ExcelExporter isolado por tenant
    """
    if tenant_id not in _exporters:
        _exporters[tenant_id] = ExcelExporter(tenant_id)
    return _exporters[tenant_id]
