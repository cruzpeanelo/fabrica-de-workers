# -*- coding: utf-8 -*-
"""
Excel Connector
===============
Conector para exportacao e importacao de dados Excel.

Funcionalidades:
- Exportar dados para Excel (.xlsx)
- Importar dados de Excel
- Templates personalizaveis
- Formatacao automatica
- Graficos embutidos
- Power Query connection strings

Issue #117 - Conectores Nativos para Power BI, Tableau e Excel
"""

import os
import io
import logging
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Union, BinaryIO
from enum import Enum

logger = logging.getLogger(__name__)

# Verificar disponibilidade de bibliotecas
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class ExcelChartType(str, Enum):
    """Tipos de grafico Excel"""
    BAR = "bar"
    PIE = "pie"
    LINE = "line"
    COLUMN = "column"
    AREA = "area"


class ExcelTableStyle(str, Enum):
    """Estilos de tabela Excel"""
    LIGHT1 = "TableStyleLight1"
    LIGHT9 = "TableStyleLight9"
    MEDIUM2 = "TableStyleMedium2"
    MEDIUM9 = "TableStyleMedium9"
    DARK1 = "TableStyleDark1"


@dataclass
class ExcelConfig:
    """Configuracao do Excel Connector"""
    default_template: str = ""
    company_name: str = "Plataforma E"
    company_color: str = "003B4A"  # Azul Belgo
    accent_color: str = "FF6C00"  # Laranja Belgo
    date_format: str = "DD/MM/YYYY"
    number_format: str = "#,##0.00"
    auto_filter: bool = True
    freeze_header: bool = True
    auto_width: bool = True

    @classmethod
    def from_env(cls) -> "ExcelConfig":
        return cls(
            default_template=os.getenv("EXCEL_TEMPLATE", ""),
            company_name=os.getenv("EXCEL_COMPANY_NAME", "Plataforma E"),
            company_color=os.getenv("EXCEL_COMPANY_COLOR", "003B4A"),
            accent_color=os.getenv("EXCEL_ACCENT_COLOR", "FF6C00")
        )


@dataclass
class ExcelColumn:
    """Definicao de coluna Excel"""
    name: str
    header: str
    width: int = 15
    format: Optional[str] = None
    align: str = "left"


@dataclass
class ExcelSheet:
    """Definicao de planilha Excel"""
    name: str
    columns: List[ExcelColumn]
    data: List[Dict]
    title: Optional[str] = None
    add_chart: bool = False
    chart_type: ExcelChartType = ExcelChartType.BAR
    table_style: ExcelTableStyle = ExcelTableStyle.MEDIUM9


class ExcelConnector:
    """
    Conector para Excel.

    Permite:
    - Criar workbooks com multiplas planilhas
    - Formatacao automatica
    - Adicionar graficos
    - Tabelas dinamicas
    - Power Query connections

    Exemplo:
    ```python
    config = ExcelConfig()
    connector = ExcelConnector(config)

    # Definir colunas
    columns = [
        ExcelColumn("story_id", "ID", width=12),
        ExcelColumn("title", "Titulo", width=40),
        ExcelColumn("status", "Status", width=15),
        ExcelColumn("story_points", "Pontos", width=10, format="#,##0")
    ]

    # Dados
    data = [
        {"story_id": "STR-001", "title": "Story 1", "status": "done", "story_points": 5},
        {"story_id": "STR-002", "title": "Story 2", "status": "in_progress", "story_points": 3}
    ]

    # Criar Excel
    excel_bytes = connector.create_workbook([
        ExcelSheet(name="Stories", columns=columns, data=data, add_chart=True)
    ])

    # Salvar
    with open("stories.xlsx", "wb") as f:
        f.write(excel_bytes)
    ```
    """

    def __init__(self, config: Optional[ExcelConfig] = None, tenant_id: str = ""):
        self.config = config or ExcelConfig()
        self.tenant_id = tenant_id

        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl nao instalado. Use: pip install openpyxl")

    def _get_header_style(self) -> Dict:
        """Retorna estilo do cabecalho"""
        return {
            "font": Font(bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=self.config.company_color,
                              end_color=self.config.company_color,
                              fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }

    def _get_cell_style(self, align: str = "left") -> Dict:
        """Retorna estilo de celula"""
        return {
            "alignment": Alignment(horizontal=align, vertical="center"),
            "border": Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC")
            )
        }

    def _apply_style(self, cell, style: Dict):
        """Aplica estilo a uma celula"""
        if "font" in style:
            cell.font = style["font"]
        if "fill" in style:
            cell.fill = style["fill"]
        if "alignment" in style:
            cell.alignment = style["alignment"]
        if "border" in style:
            cell.border = style["border"]

    def _auto_width(self, ws, columns: List[ExcelColumn]):
        """Ajusta largura das colunas"""
        for idx, col in enumerate(columns, 1):
            column_letter = openpyxl.utils.get_column_letter(idx)
            ws.column_dimensions[column_letter].width = col.width

    def _add_chart(
        self,
        ws,
        chart_type: ExcelChartType,
        data_range: str,
        title: str = "",
        position: str = "E2"
    ):
        """Adiciona grafico a planilha"""
        if chart_type == ExcelChartType.BAR:
            chart = BarChart()
        elif chart_type == ExcelChartType.PIE:
            chart = PieChart()
        elif chart_type == ExcelChartType.LINE:
            chart = LineChart()
        else:
            chart = BarChart()

        chart.title = title
        chart.style = 10

        # Dados do grafico
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 15
        chart.height = 10

        ws.add_chart(chart, position)

    def create_workbook(
        self,
        sheets: List[ExcelSheet],
        title: Optional[str] = None
    ) -> bytes:
        """
        Cria workbook Excel.

        Args:
            sheets: Lista de planilhas
            title: Titulo do documento

        Returns:
            Bytes do arquivo Excel
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl e necessario para criar Excel")

        wb = Workbook()

        # Remover planilha padrao
        if wb.active:
            wb.remove(wb.active)

        for sheet in sheets:
            ws = wb.create_sheet(title=sheet.name[:31])  # Limite do Excel

            # Titulo da planilha (opcional)
            start_row = 1
            if sheet.title:
                ws.merge_cells(f"A1:{openpyxl.utils.get_column_letter(len(sheet.columns))}1")
                title_cell = ws.cell(row=1, column=1, value=sheet.title)
                title_cell.font = Font(bold=True, size=14, color=self.config.company_color)
                title_cell.alignment = Alignment(horizontal="center")
                start_row = 3

            # Cabecalhos
            header_style = self._get_header_style()
            for col_idx, column in enumerate(sheet.columns, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=column.header)
                self._apply_style(cell, header_style)

            # Dados
            cell_style = self._get_cell_style()
            for row_idx, row_data in enumerate(sheet.data, start_row + 1):
                for col_idx, column in enumerate(sheet.columns, 1):
                    value = row_data.get(column.name, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    self._apply_style(cell, self._get_cell_style(column.align))

                    # Formato numerico
                    if column.format:
                        cell.number_format = column.format

            # Tabela formatada
            if sheet.data:
                end_row = start_row + len(sheet.data)
                end_col = openpyxl.utils.get_column_letter(len(sheet.columns))
                table_range = f"A{start_row}:{end_col}{end_row}"

                table = Table(displayName=sheet.name.replace(" ", "_"), ref=table_range)
                style = TableStyleInfo(
                    name=sheet.table_style.value,
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                ws.add_table(table)

            # Auto-width
            if self.config.auto_width:
                self._auto_width(ws, sheet.columns)

            # Freeze panes
            if self.config.freeze_header:
                ws.freeze_panes = f"A{start_row + 1}"

            # Grafico
            if sheet.add_chart and sheet.data:
                chart_position = openpyxl.utils.get_column_letter(len(sheet.columns) + 2) + str(start_row)
                self._add_chart(ws, sheet.chart_type, table_range, sheet.title or sheet.name, chart_position)

        # Salvar em bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.read()

    def create_stories_report(
        self,
        stories: List[Dict],
        include_charts: bool = True
    ) -> bytes:
        """
        Cria relatorio Excel de stories.

        Args:
            stories: Lista de stories
            include_charts: Incluir graficos

        Returns:
            Bytes do arquivo Excel
        """
        # Calcular metricas
        status_counts = {}
        priority_counts = {}
        total_points = 0

        for story in stories:
            status = story.get("status", "unknown")
            priority = story.get("priority", "unknown")
            points = story.get("story_points", 0) or 0

            status_counts[status] = status_counts.get(status, 0) + 1
            priority_counts[priority] = priority_counts.get(priority, 0) + 1
            total_points += points

        # Planilha de stories
        stories_columns = [
            ExcelColumn("story_id", "ID", width=12),
            ExcelColumn("title", "Titulo", width=45),
            ExcelColumn("status", "Status", width=15),
            ExcelColumn("priority", "Prioridade", width=12),
            ExcelColumn("story_points", "Pontos", width=10, format="#,##0", align="center"),
            ExcelColumn("assignee", "Responsavel", width=20),
            ExcelColumn("sprint_id", "Sprint", width=12),
            ExcelColumn("created_at", "Criado em", width=18)
        ]

        stories_sheet = ExcelSheet(
            name="Stories",
            title=f"Relatorio de Stories - {self.config.company_name}",
            columns=stories_columns,
            data=stories,
            table_style=ExcelTableStyle.MEDIUM9
        )

        # Planilha de resumo por status
        status_data = [{"status": k, "count": v} for k, v in status_counts.items()]
        status_columns = [
            ExcelColumn("status", "Status", width=20),
            ExcelColumn("count", "Quantidade", width=15, format="#,##0", align="center")
        ]

        status_sheet = ExcelSheet(
            name="Por Status",
            title="Resumo por Status",
            columns=status_columns,
            data=status_data,
            add_chart=include_charts,
            chart_type=ExcelChartType.PIE
        )

        # Planilha de resumo por prioridade
        priority_data = [{"priority": k, "count": v} for k, v in priority_counts.items()]
        priority_columns = [
            ExcelColumn("priority", "Prioridade", width=20),
            ExcelColumn("count", "Quantidade", width=15, format="#,##0", align="center")
        ]

        priority_sheet = ExcelSheet(
            name="Por Prioridade",
            title="Resumo por Prioridade",
            columns=priority_columns,
            data=priority_data,
            add_chart=include_charts,
            chart_type=ExcelChartType.BAR
        )

        return self.create_workbook([stories_sheet, status_sheet, priority_sheet])

    def create_sprint_report(
        self,
        sprint: Dict,
        stories: List[Dict]
    ) -> bytes:
        """
        Cria relatorio de sprint.

        Args:
            sprint: Dados do sprint
            stories: Stories do sprint

        Returns:
            Bytes do arquivo Excel
        """
        # Planilha de resumo
        summary_data = [{
            "metric": "Nome do Sprint",
            "value": sprint.get("name", "")
        }, {
            "metric": "Data Inicio",
            "value": sprint.get("start_date", "")
        }, {
            "metric": "Data Fim",
            "value": sprint.get("end_date", "")
        }, {
            "metric": "Pontos Planejados",
            "value": sprint.get("planned_points", 0)
        }, {
            "metric": "Pontos Concluidos",
            "value": sprint.get("completed_points", 0)
        }, {
            "metric": "Velocidade",
            "value": sprint.get("velocity", 0)
        }, {
            "metric": "Taxa de Conclusao",
            "value": f"{sprint.get('completion_rate', 0)}%"
        }]

        summary_columns = [
            ExcelColumn("metric", "Metrica", width=25),
            ExcelColumn("value", "Valor", width=20, align="center")
        ]

        summary_sheet = ExcelSheet(
            name="Resumo",
            title=f"Sprint Report - {sprint.get('name', '')}",
            columns=summary_columns,
            data=summary_data
        )

        # Planilha de stories do sprint
        stories_columns = [
            ExcelColumn("story_id", "ID", width=12),
            ExcelColumn("title", "Titulo", width=40),
            ExcelColumn("status", "Status", width=15),
            ExcelColumn("story_points", "Pontos", width=10, format="#,##0", align="center"),
            ExcelColumn("assignee", "Responsavel", width=20)
        ]

        stories_sheet = ExcelSheet(
            name="Stories",
            columns=stories_columns,
            data=stories
        )

        return self.create_workbook([summary_sheet, stories_sheet])

    def read_excel(
        self,
        file: Union[str, BinaryIO],
        sheet_name: Optional[str] = None
    ) -> List[Dict]:
        """
        Le dados de arquivo Excel.

        Args:
            file: Caminho ou file-like object
            sheet_name: Nome da planilha (None = primeira)

        Returns:
            Lista de dicionarios com dados
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl e necessario")

        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)

        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        data = []

        for row in rows[1:]:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            data.append(row_dict)

        return data

    def get_power_query_connection(
        self,
        base_url: str,
        endpoint: str,
        api_key: Optional[str] = None
    ) -> str:
        """
        Gera connection string para Power Query.

        Args:
            base_url: URL base da API
            endpoint: Endpoint de dados
            api_key: Chave da API (opcional)

        Returns:
            String de conexao M/Power Query
        """
        connection = f'''let
    Source = Json.Document(Web.Contents("{base_url}", [
        RelativePath = "{endpoint}",
        Headers = [
            Accept = "application/json"'''

        if api_key:
            connection += f''',
            Authorization = "Bearer {api_key}"'''

        if self.tenant_id:
            connection += f''',
            #"X-Tenant-ID" = "{self.tenant_id}"'''

        connection += '''
        ]
    ])),
    #"Converted to Table" = Table.FromRecords(Source[data])
in
    #"Converted to Table"'''

        return connection


# Singleton
_excel_instance: Optional[ExcelConnector] = None


def get_excel_connector(tenant_id: str = "") -> ExcelConnector:
    """Retorna instancia global"""
    global _excel_instance
    if _excel_instance is None:
        _excel_instance = ExcelConnector(tenant_id=tenant_id)
    return _excel_instance
