# -*- coding: utf-8 -*-
"""
Report Exporters - Plataforma E
===============================

Export reports to various formats (HTML, CSV, JSON, PDF, Excel).

Issue #446: Report Generator - Geracao de Relatorios
"""

from abc import ABC, abstractmethod
from datetime import datetime
from typing import Any, Dict, List, TYPE_CHECKING
import csv
import io
import json

if TYPE_CHECKING:
    from .report_generator import ReportData, ReportFormat


class BaseExporter(ABC):
    """Base class for report exporters."""

    @abstractmethod
    def export(self, report: "ReportData") -> str:
        """Export report to string."""
        pass

    def save(self, report: "ReportData", filepath: str) -> None:
        """Save report to file."""
        content = self.export(report)
        mode = "w" if isinstance(content, str) else "wb"
        with open(filepath, mode, encoding="utf-8" if mode == "w" else None) as f:
            f.write(content)


class HTMLExporter(BaseExporter):
    """Export reports to HTML format."""

    def export(self, report: "ReportData") -> str:
        """Generate HTML report."""
        html_parts = [
            "<!DOCTYPE html>",
            "<html lang='pt-BR'>",
            "<head>",
            "  <meta charset='UTF-8'>",
            f"  <title>{report.config.title}</title>",
            "  <style>",
            self._get_styles(),
            "  </style>",
            "</head>",
            "<body>",
            f"  <header><h1>{report.config.title}</h1></header>",
            f"  <p class='meta'>Gerado em: {report.generated_at.strftime('%d/%m/%Y %H:%M')}</p>",
            f"  <p class='meta'>Tipo: {report.config.report_type.value}</p>",
        ]

        for section in sorted(report.sections, key=lambda s: s.order):
            html_parts.append(self._render_section(section))

        html_parts.extend([
            "</body>",
            "</html>",
        ])

        return "\n".join(html_parts)

    def _get_styles(self) -> str:
        """Get CSS styles for HTML report."""
        return """
        body { font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }
        header { background: #003B4A; color: white; padding: 20px; margin: -40px -40px 20px; }
        h1 { margin: 0; }
        h2 { color: #003B4A; border-bottom: 2px solid #FF6C00; padding-bottom: 5px; }
        .meta { color: #666; font-size: 0.9em; }
        .section { background: white; padding: 20px; margin: 20px 0; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        table { width: 100%; border-collapse: collapse; margin: 10px 0; }
        th { background: #003B4A; color: white; padding: 10px; text-align: left; }
        td { padding: 10px; border-bottom: 1px solid #ddd; }
        tr:nth-child(even) { background: #f9f9f9; }
        .summary { background: #e8f4f8; padding: 15px; border-radius: 5px; }
        .kpi { display: inline-block; background: #FF6C00; color: white; padding: 10px 20px; margin: 5px; border-radius: 5px; }
        """

    def _render_section(self, section) -> str:
        """Render a single section to HTML."""
        html = [f"  <div class='section'>", f"    <h2>{section.title}</h2>"]

        if section.section_type == "table" and isinstance(section.content, list):
            html.append(self._render_table(section.content))
        elif section.section_type == "summary":
            html.append(self._render_summary(section.content))
        elif isinstance(section.content, dict):
            html.append(self._render_dict(section.content))
        elif isinstance(section.content, list):
            html.append(self._render_list(section.content))
        else:
            html.append(f"    <p>{section.content}</p>")

        html.append("  </div>")
        return "\n".join(html)

    def _render_table(self, items: List[Dict]) -> str:
        """Render a list of dicts as HTML table."""
        if not items:
            return "<p>No data</p>"

        headers = list(items[0].keys())
        rows = ["    <table>", "      <tr>"]
        rows.extend([f"        <th>{h.replace('_', ' ').title()}</th>" for h in headers])
        rows.append("      </tr>")

        for item in items:
            rows.append("      <tr>")
            for h in headers:
                rows.append(f"        <td>{item.get(h, '')}</td>")
            rows.append("      </tr>")

        rows.append("    </table>")
        return "\n".join(rows)

    def _render_dict(self, data: Dict) -> str:
        """Render a dict as key-value pairs."""
        rows = ["    <dl>"]
        for key, value in data.items():
            if isinstance(value, dict):
                rows.append(f"      <dt><strong>{key.replace('_', ' ').title()}</strong></dt>")
                rows.append("      <dd>")
                rows.append(self._render_dict(value))
                rows.append("      </dd>")
            elif isinstance(value, list):
                rows.append(f"      <dt><strong>{key.replace('_', ' ').title()}</strong></dt>")
                rows.append("      <dd>")
                rows.append(self._render_list(value))
                rows.append("      </dd>")
            else:
                rows.append(f"      <dt>{key.replace('_', ' ').title()}</dt>")
                rows.append(f"      <dd>{value}</dd>")
        rows.append("    </dl>")
        return "\n".join(rows)

    def _render_list(self, items: List) -> str:
        """Render a list as HTML."""
        if not items:
            return "<p>No items</p>"

        if isinstance(items[0], dict):
            return self._render_table(items)

        rows = ["    <ul>"]
        for item in items:
            rows.append(f"      <li>{item}</li>")
        rows.append("    </ul>")
        return "\n".join(rows)

    def _render_summary(self, data: Dict) -> str:
        """Render summary section."""
        rows = ["    <div class='summary'>"]
        for key, value in data.items():
            rows.append(f"      <span class='kpi'>{key.replace('_', ' ').title()}: {value}</span>")
        rows.append("    </div>")
        return "\n".join(rows)


class CSVExporter(BaseExporter):
    """Export reports to CSV format."""

    def export(self, report: "ReportData") -> str:
        """Generate CSV report."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header info
        writer.writerow(["Report", report.config.title])
        writer.writerow(["Type", report.config.report_type.value])
        writer.writerow(["Generated", report.generated_at.isoformat()])
        writer.writerow([])

        for section in sorted(report.sections, key=lambda s: s.order):
            writer.writerow([f"=== {section.title} ==="])

            if isinstance(section.content, list) and section.content:
                if isinstance(section.content[0], dict):
                    headers = list(section.content[0].keys())
                    writer.writerow(headers)
                    for item in section.content:
                        writer.writerow([item.get(h, "") for h in headers])
                else:
                    for item in section.content:
                        writer.writerow([item])
            elif isinstance(section.content, dict):
                self._write_dict(writer, section.content)
            else:
                writer.writerow([section.content])

            writer.writerow([])

        return output.getvalue()

    def _write_dict(self, writer, data: Dict, prefix: str = "") -> None:
        """Write dict to CSV."""
        for key, value in data.items():
            full_key = f"{prefix}.{key}" if prefix else key
            if isinstance(value, dict):
                self._write_dict(writer, value, full_key)
            elif isinstance(value, list):
                writer.writerow([full_key, json.dumps(value)])
            else:
                writer.writerow([full_key, value])


class JSONExporter(BaseExporter):
    """Export reports to JSON format."""

    def export(self, report: "ReportData") -> str:
        """Generate JSON report."""
        return json.dumps(report.to_dict(), indent=2, ensure_ascii=False, default=str)


class PDFExporter(BaseExporter):
    """Export reports to PDF format.

    Note: Requires weasyprint or reportlab for actual PDF generation.
    This is a stub that generates HTML with PDF-ready styling.
    """

    def export(self, report: "ReportData") -> str:
        """Generate PDF-ready HTML (requires weasyprint to convert)."""
        html_exporter = HTMLExporter()
        html_content = html_exporter.export(report)

        # Add print-friendly styles
        print_styles = """
        @media print {
            body { margin: 0; padding: 20px; }
            .section { page-break-inside: avoid; }
        }
        @page { size: A4; margin: 2cm; }
        """

        return html_content.replace("</style>", print_styles + "</style>")

    def save_as_pdf(self, report: "ReportData", filepath: str) -> bool:
        """Save as actual PDF (requires weasyprint)."""
        try:
            from weasyprint import HTML
            html_content = self.export(report)
            HTML(string=html_content).write_pdf(filepath)
            return True
        except ImportError:
            # Fallback: save as HTML
            html_path = filepath.replace(".pdf", ".html")
            self.save(report, html_path)
            return False


class ExcelExporter(BaseExporter):
    """Export reports to Excel format.

    Note: Requires openpyxl for actual Excel generation.
    This is a stub that generates CSV as fallback.
    """

    def export(self, report: "ReportData") -> str:
        """Generate Excel-compatible CSV."""
        csv_exporter = CSVExporter()
        return csv_exporter.export(report)

    def save_as_xlsx(self, report: "ReportData", filepath: str) -> bool:
        """Save as actual Excel file (requires openpyxl)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = report.config.report_type.value[:31]

            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="003B4A")

            row = 1
            ws.cell(row=row, column=1, value="Report").font = Font(bold=True)
            ws.cell(row=row, column=2, value=report.config.title)
            row += 1
            ws.cell(row=row, column=1, value="Generated").font = Font(bold=True)
            ws.cell(row=row, column=2, value=report.generated_at.isoformat())
            row += 2

            for section in sorted(report.sections, key=lambda s: s.order):
                ws.cell(row=row, column=1, value=section.title).font = Font(bold=True, size=14)
                row += 1

                if isinstance(section.content, list) and section.content:
                    if isinstance(section.content[0], dict):
                        headers = list(section.content[0].keys())
                        for col, h in enumerate(headers, 1):
                            cell = ws.cell(row=row, column=col, value=h.replace("_", " ").title())
                            cell.font = header_font
                            cell.fill = header_fill
                        row += 1

                        for item in section.content:
                            for col, h in enumerate(headers, 1):
                                ws.cell(row=row, column=col, value=item.get(h, ""))
                            row += 1
                elif isinstance(section.content, dict):
                    for key, value in section.content.items():
                        ws.cell(row=row, column=1, value=key.replace("_", " ").title())
                        ws.cell(row=row, column=2, value=str(value) if not isinstance(value, (dict, list)) else json.dumps(value))
                        row += 1

                row += 1

            wb.save(filepath)
            return True
        except ImportError:
            # Fallback: save as CSV
            csv_path = filepath.replace(".xlsx", ".csv")
            csv_exporter = CSVExporter()
            csv_exporter.save(report, csv_path)
            return False


class ExporterFactory:
    """Factory for creating report exporters."""

    _exporters = {
        "html": HTMLExporter,
        "csv": CSVExporter,
        "json": JSONExporter,
        "pdf": PDFExporter,
        "excel": ExcelExporter,
    }

    @classmethod
    def create(cls, format: "ReportFormat") -> BaseExporter:
        """Create an exporter for the given format."""
        format_str = format.value if hasattr(format, "value") else str(format)
        exporter_class = cls._exporters.get(format_str.lower())
        if not exporter_class:
            raise ValueError(f"Unknown format: {format}")
        return exporter_class()

    @classmethod
    def register(cls, format_name: str, exporter_class: type) -> None:
        """Register a custom exporter."""
        cls._exporters[format_name.lower()] = exporter_class

    @classmethod
    def available_formats(cls) -> List[str]:
        """Get list of available export formats."""
        return list(cls._exporters.keys())
